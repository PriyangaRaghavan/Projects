## Add a new column to classify examples as prestate and poststate

import pandas as pd
from openpyxl import workbook 
from openpyxl import load_workbook

x1 = pd.ExcelFile("identify_action.xlsx")

## load workbook
workbook = load_workbook('identify_action.xlsx')
sheets = workbook.sheetnames
if sheets[0] == "Sheet1":
   del workbook['Sheet1']
   workbook.save('identify_action.xlsx')

sheets = workbook.sheetnames
num_sheet = len(sheets)


### func def to create new column "Class"
def write_class(sheet, column_count):
    sheet = sheet
    #column_count= sheet.max_column
    sheet.cell(row = 1, column = column_count+1).value = "Class"
    workbook.save("identify_action.xlsx")

### func def to classify prestate and poststate
def class_value(sheet, num_rows,column_count):
    sheet = sheet
    #column_count= sheet.max_column
    for j in range(2, num_rows+1, 2):
        sheet.cell(row = j, column = column_count).value = "prestate"
        sheet.cell(row = j+1, column = column_count).value = "poststate"
    workbook.save("identify_action.xlsx")

e =0
# function call to create new column "Class"
for j in range (num_sheet):
    sheet = workbook[sheets[j]]
    df = x1.parse(sheets[j])
    num_rows = len(df)    
    columns = df.columns
    column_count= sheet.max_column
    for k in range (column_count):
        if columns[k] == "Class":
           e =1
           break
    if e !=1:
        write_class(sheet, column_count)
    
    column_count= sheet.max_column
    class_value(sheet,num_rows,column_count)

print("Prestate and Poststate of actions")